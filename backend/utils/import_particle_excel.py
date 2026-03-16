"""
Particle 과거 데이터 Excel 일괄 업로드 스크립트

사용법:
    1. 아래 '설정' 섹션에서 EXCEL_FILE_PATH와 COLUMN_MAP을 실제 파일에 맞게 수정
    2. 가상환경 활성화 후 실행:
       cd "C:\\Program Files\\DICD_Management_System"
       call venv\\Scripts\\activate.bat
       python backend\\utils\\import_particle_excel.py
"""
import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from datetime import datetime
from openpyxl import load_workbook
from backend.database.database import SessionLocal
from backend.database.models import ParticleMeasurement, ParticleEquipment
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)


# ============================================================
# 설정
# ============================================================

# Excel 파일 경로
EXCEL_FILE_PATH = "particle_data.xlsx"

# 시트 이름 (None이면 첫 번째 시트 사용)
SHEET_NAME = None

# 데이터가 시작하는 행 번호 (헤더 다음 행)
DATA_START_ROW = 2

# Excel 컬럼 → DB 필드 매핑 (실제 Excel 헤더에 맞게 설정됨)
COLUMN_MAP = {
    "장비번호":   "equipment_name",     # 장비 이름으로 DB 조회
    "측정전Y":    "before_y",
    "측정전O":    "before_o",
    "측정전B":    "before_b",
    "측정후Y":    "after_y",
    "측정후O":    "after_o",
    "측정후B":    "after_b",
    "날짜":       "created_at",
    "작성자":     "author",
}

# 장비 조회 방식: "equipment_name" (이름으로 조회) 또는 "equipment_number" (번호로 조회)
EQUIPMENT_FIELD = "equipment_name"

# Excel 장비명 → DB 장비명 매핑 (Excel 이름과 DB 이름이 다를 경우 여기에 추가)
# 예: "Coater 1호기 1라인_#08": "장비1"
EQUIPMENT_NAME_MAP = {
    # "Excel에서의 장비명": "DB에 등록된 장비명",
}

# 작성자 기본값 (Excel에 작성자 컬럼이 없는 경우)
DEFAULT_AUTHOR = "데이터 마이그레이션"

# ============================================================
# 이하 수정 불필요
# ============================================================


def get_equipment_id(db, value, field="equipment_number"):
    """장비 번호 또는 이름으로 DB equipment_id 조회"""
    if value is None:
        return None, "장비 값이 비어있음"

    if field == "equipment_number":
        try:
            num = int(value)
        except (ValueError, TypeError):
            return None, f"장비번호 변환 실패: {value}"
        eq = db.query(ParticleEquipment).filter(ParticleEquipment.equipment_number == num).first()
    else:
        # EQUIPMENT_NAME_MAP으로 이름 변환 후 조회
        name = EQUIPMENT_NAME_MAP.get(str(value), str(value))
        eq = db.query(ParticleEquipment).filter(ParticleEquipment.name == name).first()

    if not eq:
        return None, f"장비를 찾을 수 없음: '{value}' (DB에 해당 이름의 장비가 없습니다)"
    return eq.id, None


def parse_datetime(value):
    """다양한 형식의 날짜/시간 값을 datetime으로 변환"""
    if isinstance(value, datetime):
        return value, None
    if value is None:
        return datetime.now(), None

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%d/%m/%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(str(value).strip(), fmt), None
    # for-else: 모든 형식 실패
    return None, f"날짜 파싱 실패: {value}"


def safe_int(value, field_name):
    """정수 변환 (None/빈값은 0 반환)"""
    if value is None or str(value).strip() == "":
        return 0, None
    try:
        return int(float(str(value))), None
    except (ValueError, TypeError):
        return None, f"{field_name} 정수 변환 실패: {value}"


def run_import():
    if not os.path.exists(EXCEL_FILE_PATH):
        logger.error(f"파일을 찾을 수 없습니다: {EXCEL_FILE_PATH}")
        sys.exit(1)

    logger.info(f"Excel 파일 로딩: {EXCEL_FILE_PATH}")
    wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
    logger.info(f"시트: {ws.title}")

    # 헤더 행 읽기
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info(f"헤더: {headers}")

    # 헤더 → 컬럼 인덱스 매핑
    header_index = {h: i for i, h in enumerate(headers) if h is not None}

    # COLUMN_MAP 검증
    missing = []
    for excel_col, db_field in COLUMN_MAP.items():
        if db_field in ("created_at", "author"):
            continue  # 선택 필드
        if excel_col not in header_index:
            missing.append(excel_col)
    if missing:
        logger.warning(f"다음 컬럼이 Excel에 없습니다 (0으로 처리됨): {missing}")

    def get_cell(row_values, excel_col):
        idx = header_index.get(excel_col)
        return row_values[idx] if idx is not None and idx < len(row_values) else None

    def db_field_col(db_field):
        """db_field에 해당하는 excel 컬럼명 반환"""
        for k, v in COLUMN_MAP.items():
            if v == db_field:
                return k
        return None

    db = SessionLocal()
    success_count = 0
    fail_count = 0
    errors = []

    try:
        rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
        total = len([r for r in rows if any(c is not None for c in r)])
        logger.info(f"총 {total}개 행 처리 시작\n")

        batch = []
        BATCH_SIZE = 500

        for row_num, row in enumerate(rows, start=DATA_START_ROW):
            if all(c is None for c in row):
                continue

            row_errors = []

            # 장비 ID
            eq_excel_col = db_field_col(EQUIPMENT_FIELD)
            eq_raw = get_cell(row, eq_excel_col) if eq_excel_col else None
            equipment_id, err = get_equipment_id(db, eq_raw, EQUIPMENT_FIELD)
            if err:
                row_errors.append(err)

            # before 값
            before_y, err = safe_int(get_cell(row, db_field_col("before_y")), "before_y")
            if err: row_errors.append(err)
            before_o, err = safe_int(get_cell(row, db_field_col("before_o")), "before_o")
            if err: row_errors.append(err)
            before_b, err = safe_int(get_cell(row, db_field_col("before_b")), "before_b")
            if err: row_errors.append(err)

            # after 값
            after_y, err = safe_int(get_cell(row, db_field_col("after_y")), "after_y")
            if err: row_errors.append(err)
            after_o, err = safe_int(get_cell(row, db_field_col("after_o")), "after_o")
            if err: row_errors.append(err)
            after_b, err = safe_int(get_cell(row, db_field_col("after_b")), "after_b")
            if err: row_errors.append(err)

            # 날짜
            dt_col = db_field_col("created_at")
            created_at, err = parse_datetime(get_cell(row, dt_col) if dt_col else None)
            if err: row_errors.append(err)

            # 작성자
            author_col = db_field_col("author")
            author = get_cell(row, author_col) if author_col else None
            author = str(author).strip() if author else DEFAULT_AUTHOR

            if row_errors:
                fail_count += 1
                errors.append(f"행 {row_num}: {'; '.join(row_errors)}")
                continue

            # final 값 계산
            final_y = max(0, (after_y or 0) - (before_y or 0))
            final_o = max(0, (after_o or 0) - (before_o or 0))
            final_b = max(0, (after_b or 0) - (before_b or 0))
            value = final_y + final_o + final_b

            batch.append(ParticleMeasurement(
                equipment_id=equipment_id,
                before_y=before_y,
                before_o=before_o,
                before_b=before_b,
                after_y=after_y,
                after_o=after_o,
                after_b=after_b,
                final_y=final_y,
                final_o=final_o,
                final_b=final_b,
                value=value,
                author=author,
                created_at=created_at,
            ))
            success_count += 1

            # 배치 저장
            if len(batch) >= BATCH_SIZE:
                db.bulk_save_objects(batch)
                db.commit()
                logger.info(f"  {success_count}건 저장 완료...")
                batch = []

        # 나머지 저장
        if batch:
            db.bulk_save_objects(batch)
            db.commit()

        logger.info(f"\n✅ 완료: 성공 {success_count}건 / 실패 {fail_count}건")

        if errors:
            logger.warning(f"\n실패 목록 (최대 20건):")
            for e in errors[:20]:
                logger.warning(f"  - {e}")
            if len(errors) > 20:
                logger.warning(f"  ... 외 {len(errors) - 20}건")

    except Exception as e:
        db.rollback()
        logger.error(f"❌ 오류 발생: {e}")
        raise
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("Particle 과거 데이터 Excel 업로드")
    logger.info("=" * 60)
    run_import()
